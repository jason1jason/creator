import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pymysql
from openpyxl import Workbook, load_workbook
from datetime import datetime
import pandas as pd
import sys
import os

class DarenApp:
    def __init__(self, root):
        self.root = root
        self.root.title("达人管理系统 v2.0")
        self.root.geometry("1200x800")
        
        # ==== 新增的路径处理代码开始 ====
        # 获取程序所在目录
        if getattr(sys, 'frozen', False):
            # 打包后的情况 - 使用exe所在目录
            application_path = os.path.dirname(sys.executable)
        else:
            # 开发时的情况 - 使用脚本所在目录
            application_path = os.path.dirname(os.path.abspath(__file__))
        
        # 构建数据库完整路径
        db_path = os.path.join(application_path, 'daren.db')
        # ==== 新增的路径处理代码结束 ====
        
        # 初始化数据库连接（修改这行）
        self.conn = pymysql.connect(
        host="192.168.0.111",  # 你的局域网 IP 地址
        user="admin",  # MySQL 用户名
        password="123456",  # MySQL 密码
        database="talent_management",  # MySQL 数据库名称
        charset="utf8mb4")
        self.c = self.conn.cursor()
        
        # 控制变量
        self.merge_duplicates = tk.BooleanVar(value=True)
        
        # 创建界面组件
        self.create_widgets()

    def create_widgets(self):
        """创建所有界面组件"""
        # 顶部标题
        ttk.Label(self.root, text="达人信息管理系统", font=('Arial', 16)).pack(pady=10)
        
        # 输入区域框架
        input_frame = ttk.LabelFrame(self.root, text="达人信息录入")
        input_frame.pack(padx=10, pady=5, fill="x")
        
        # 名称输入
        ttk.Label(input_frame, text="达人名称:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.name_entry = ttk.Entry(input_frame, width=30)
        self.name_entry.grid(row=0, column=1, padx=5, pady=5)
        
        # 店铺输入
        ttk.Label(input_frame, text="店铺名称:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.shop_entry = ttk.Entry(input_frame, width=30)
        self.shop_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # 销售额输入
        ttk.Label(input_frame, text="销售额:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.sales_entry = ttk.Entry(input_frame, width=30)
        self.sales_entry.grid(row=2, column=1, padx=5, pady=5)
        
        # 黑名单复选框
        self.is_black = tk.BooleanVar()
        ttk.Checkbutton(input_frame, text="加入黑名单", variable=self.is_black).grid(row=3, column=1, sticky="w", padx=5)
        
        # 样品订单号输入
        ttk.Label(input_frame, text="样品订单号:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        self.sample_order_entry = ttk.Entry(input_frame, width=30)
        self.sample_order_entry.grid(row=4, column=1, padx=5, pady=5)

        # 产品 SKU 名称输入
        ttk.Label(input_frame, text="产品 SKU 名称:").grid(row=5, column=0, padx=5, pady=5, sticky="e")
        self.sku_name_entry = ttk.Entry(input_frame, width=30)
        self.sku_name_entry.grid(row=5, column=1, padx=5, pady=5)

        # 是否创作视频复选框
        self.has_video = tk.BooleanVar()
        ttk.Checkbutton(input_frame, text="是否创作视频", variable=self.has_video).grid(row=6, column=1, sticky="w", padx=5)

        # 备注输入 (调整到第7行)
        ttk.Label(input_frame, text="备注:").grid(row=7, column=0, padx=5, pady=5, sticky="e")
        self.remark_entry = ttk.Entry(input_frame, width=30)
        self.remark_entry.grid(row=7, column=1, padx=5, pady=5)

        # 按钮框架 (调整到第8行)
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=8, column=0, columnspan=2, pady=10, sticky="ew")
        
        # 均匀分布按钮
        ttk.Button(button_frame, text="添加记录", command=self.add_daren).pack(side="left", expand=True)
        ttk.Button(button_frame, text="导入数据", command=self.import_data).pack(side="left", expand=True)
        ttk.Button(button_frame, text="导出模板", command=self.export_template).pack(side="left", expand=True)

        
        # 搜索区域框架
        search_frame = ttk.LabelFrame(self.root, text="达人查询")
        search_frame.pack(padx=10, pady=5, fill="both", expand=True)
        
        # 搜索控制区域
        control_frame = ttk.Frame(search_frame)
        control_frame.pack(fill="x", pady=5)
        
        # 搜索框
        ttk.Label(control_frame, text="搜索名称:").pack(side="left", padx=5)
        self.search_entry = ttk.Entry(control_frame, width=30)
        self.search_entry.pack(side="left", padx=5)
        
        # 修改搜索框标签
        ttk.Label(control_frame, text="批量搜索(逗号/换行分隔):").pack(side="left", padx=5)

        # 搜索按钮
        ttk.Button(control_frame, text="查询", command=self.search_daren).pack(side="left", padx=5)
        
        # 合并选项
        ttk.Checkbutton(
            control_frame, 
            text="合并同名达人", 
            variable=self.merge_duplicates,
            command=self.search_daren  # 切换时自动刷新查询
        ).pack(side="left", padx=5)
        
        # 操作按钮
        ttk.Button(control_frame, text="导出数据", command=self.export_to_excel).pack(side="left", padx=5)
        ttk.Button(control_frame, text="删除选中", command=self.delete_selected).pack(side="left", padx=5)
        #修改按钮
        ttk.Button(control_frame, text="修改选中", command=self.edit_selected).pack(side="left", padx=5)
        # 结果显示表格
        self.tree = ttk.Treeview(
            search_frame, 
            columns=("id", "name", "shops", "sales", "black", "sample_order", "sku_name", "has_video", "remark"), 
            show="headings",
            selectmode="extended"  # 允许多选
        )
        
        # 设置表头
        self.tree.heading("id", text="ID")
        self.tree.heading("name", text="达人名称")
        self.tree.heading("shops", text="所属店铺")
        self.tree.heading("sales", text="总销售额")
        self.tree.heading("black", text="黑名单")
        self.tree.heading("sample_order", text="样品订单号")
        self.tree.heading("sku_name", text="产品 SKU 名称")
        self.tree.heading("has_video", text="是否创作视频")
        self.tree.heading("remark", text="备注")
        
        # 设置列宽
        self.tree.column("id", width=40, anchor="center")
        self.tree.column("name", width=120)
        self.tree.column("shops", width=150)
        self.tree.column("sales", width=80, anchor="e")
        self.tree.column("black", width=60, anchor="center")
        self.tree.column("sample_order", width=120)
        self.tree.column("sku_name", width=150)
        self.tree.column("has_video", width=100, anchor="center")
        self.tree.column("remark", width=150)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(search_frame, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 初始化表格数据
        self.search_daren()

    def edit_selected(self):
    
        """编辑选中的记录"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("警告", "请先选择要修改的记录")
            return

        # 获取选中记录的值
        item = self.tree.item(selected_items[0])
        values = item["values"]

        # 弹出修改窗口 - 增大对话框尺寸
        edit_dialog = tk.Toplevel(self.root)
        edit_dialog.title("修改记录")
        edit_dialog.geometry("500x500")  # 增大对话框尺寸
        edit_dialog.grab_set()

        # 创建输入框 - 使用grid布局并调整行号
        row = 0
        
        # 达人名称
        ttk.Label(edit_dialog, text="达人名称:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        name_entry = ttk.Entry(edit_dialog, width=40)
        name_entry.grid(row=row, column=1, padx=10, pady=5, sticky="w")
        name_entry.insert(0, values[1])
        row += 1

        # 店铺名称
        ttk.Label(edit_dialog, text="店铺名称:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        shop_entry = ttk.Entry(edit_dialog, width=40)
        shop_entry.grid(row=row, column=1, padx=10, pady=5, sticky="w")
        shop_entry.insert(0, values[2])
        row += 1

        # 销售额
        ttk.Label(edit_dialog, text="销售额:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        sales_entry = ttk.Entry(edit_dialog, width=40)
        sales_entry.grid(row=row, column=1, padx=10, pady=5, sticky="w")
        sales_entry.insert(0, values[3])
        row += 1

        # 样品订单号
        ttk.Label(edit_dialog, text="样品订单号:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        sample_order_entry = ttk.Entry(edit_dialog, width=40)
        sample_order_entry.grid(row=row, column=1, padx=10, pady=5, sticky="w")
        sample_order_entry.insert(0, values[5])
        row += 1

        # 产品SKU名称
        ttk.Label(edit_dialog, text="产品SKU名称:").grid(row=row, column=0, padx=10, pady=5, sticky="e")
        sku_name_entry = ttk.Entry(edit_dialog, width=40)
        sku_name_entry.grid(row=row, column=1, padx=10, pady=5, sticky="w")
        sku_name_entry.insert(0, values[6])
        row += 1

        # 是否创作视频
        has_video_var = tk.BooleanVar(value=(values[7] == "是"))
        ttk.Checkbutton(edit_dialog, text="是否创作视频", variable=has_video_var).grid(
            row=row, column=1, padx=10, pady=5, sticky="w")
        row += 1

        # 备注 - 使用Text控件代替Entry以获得多行输入
        ttk.Label(edit_dialog, text="备注:").grid(row=row, column=0, padx=10, pady=5, sticky="ne")
        remark_text = tk.Text(edit_dialog, width=30, height=4, wrap=tk.WORD)
        remark_text.grid(row=row, column=1, padx=10, pady=5, sticky="w")
        remark_text.insert("1.0", values[8] if len(values) > 8 else "")
        row += 1

        # 按钮框架
        button_frame = ttk.Frame(edit_dialog)
        button_frame.grid(row=row, column=0, columnspan=2, pady=10)
        
        # 确定按钮
        ttk.Button(button_frame, text="确定", command=lambda: self.update_record(
            values[0], name_entry.get(), shop_entry.get(), sales_entry.get(),
            sample_order_entry.get(), sku_name_entry.get(), has_video_var.get(),
            remark_text.get("1.0", tk.END).strip(), edit_dialog
        )).pack(side="left", padx=10)
        
        # 取消按钮
        ttk.Button(button_frame, text="取消", command=edit_dialog.destroy).pack(side="left", padx=10)

    def batch_search_daren(self):
        """批量查询达人信息"""
        # 清空现有表格数据
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        # 获取搜索文本并处理
        search_text = self.search_entry.get("1.0", tk.END).strip()
        if not search_text:
            messagebox.showwarning("提示", "请输入要查询的达人名称")
            return
        
        # 分割输入内容（支持逗号和换行分隔）
        names = []
        for line in search_text.split('\n'):
            names.extend([name.strip() for name in line.split(',') if name.strip()])
        
        if not names:
            messagebox.showwarning("提示", "未识别到有效的达人名称")
            return
        
        # 构建SQL查询条件
        placeholders = ','.join(['%s'] * len(names))
        query = f"SELECT * FROM daren WHERE name IN ({placeholders})"
        
        try:
            self.c.execute(query, names)
            results = self.c.fetchall()
            
            if not results:
                messagebox.showinfo("结果", "未找到匹配的达人记录")
                return
            
            # 显示结果
            for row in results:
                daren_id, name, shop, sales, is_black, sample_order, sku_name, has_video, remark = row
                black_status = "是" if is_black else "否"
                video_status = "是" if has_video else "否"
                self.tree.insert("", tk.END, values=(
                    daren_id,
                    name,
                    shop,
                    sales,
                    black_status,
                    sample_order,
                    sku_name,
                    video_status,
                    remark
                ))
            
            messagebox.showinfo("完成", f"共找到 {len(results)} 条匹配记录")
        
        except Exception as e:
            messagebox.showerror("错误", f"查询过程中出错:\n{str(e)}")

    def update_record(self, record_id, name, shop, sales, sample_order, sku_name, has_video, remark, dialog):
        """更新选中的记录"""
        try:
            sales_val = float(sales) if sales else 0.0
            self.c.execute(
                """UPDATE daren
                SET name = %s, shop = %s, sales = %s, sample_order = %s, 
                    sku_name = %s,  has_video = %s,  remark = %s,
                WHERE id = %s,""",
                (name, shop, sales_val, sample_order, sku_name, 
                1 if has_video else 0, remark, record_id)
            )
            self.conn.commit()
            messagebox.showinfo("成功", "记录已更新！")
            dialog.destroy()
            self.search_daren()  # 刷新表格
        except ValueError:
            messagebox.showerror("错误", "销售额必须是数字！")
        except Exception as e:
            messagebox.showerror("错误", f"更新记录时出错：{e}")

    def add_daren(self):
        """添加新的达人记录"""
        name = self.name_entry.get()
        shop = self.shop_entry.get()
        sales = self.sales_entry.get()
        sample_order = self.sample_order_entry.get()
        sku_name = self.sku_name_entry.get()
        has_video = 1 if self.has_video.get() else 0
        remark = self.remark_entry.get()

        if not name:
            messagebox.showerror("错误", "达人名称不能为空！")
            return

        try:
            sales_val = float(sales) if sales else 0.0
            self.c.execute(
                """INSERT INTO daren (name, shop, sales, is_black, sample_order, sku_name, has_video, remark)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                (name, shop, sales_val, 1 if self.is_black.get() else 0, sample_order, sku_name, has_video, remark)
            )
            self.conn.commit()

            messagebox.showinfo("成功", "达人信息已添加！")
            # 清空输入框
            self.name_entry.delete(0, tk.END)
            self.shop_entry.delete(0, tk.END)
            self.sales_entry.delete(0, tk.END)
            self.sample_order_entry.delete(0, tk.END)
            self.sku_name_entry.delete(0, tk.END)
            self.remark_entry.delete(0, tk.END)
            self.is_black.set(False)
            self.has_video.set(False)

            # 强制刷新表格 - 先清空搜索条件
            self.search_entry.delete(0, tk.END)
            self.search_daren()
        except ValueError:
            messagebox.showerror("错误", "销售额必须是数字！")
        except Exception as e:
            messagebox.showerror("错误", f"添加记录失败: {str(e)}")
    
    def search_daren(self):
        """查询达人数据，并根据设置合并重复项"""
        # 清空现有表格数据
        for row in self.tree.get_children():
            self.tree.delete(row)
        
        # 执行查询
        name = self.search_entry.get()
        if name:
            self.c.execute("SELECT * FROM daren WHERE name LIKE %s,", ('%'+name+'%',))
        else:
            self.c.execute("SELECT * FROM daren")
        
        if self.merge_duplicates.get():
            # 合并重复名称数据
            merged_data = {}
            for row in self.c.fetchall():
                # 确保有9个值
                row = list(row)
                while len(row) < 9:
                    row.append('' if len(row) < 6 else 0)
                
                daren_id, name, shop, sales, is_black, sample_order, sku_name, has_video, remark = row
                
                if name not in merged_data:
                    merged_data[name] = {
                        'id': daren_id,
                        'shops': [shop],
                        'total_sales': sales,
                        'is_black': is_black,
                        'sample_order': sample_order,
                        'sku_name': sku_name,
                        'has_video': has_video,
                        'remark': remark
                    }
                else:
                    merged_data[name]['shops'].append(shop)
                    merged_data[name]['total_sales'] += sales
                    if is_black:
                        merged_data[name]['is_black'] = 1
            
            # 填充表格
            for name, data in merged_data.items():
                shops = ", ".join(data['shops'])
                black_status = "是" if data['is_black'] else "否"
                video_status = "是" if data['has_video'] else "否"
                self.tree.insert("", tk.END, values=(
                    data['id'],
                    name,
                    shops,
                    data['total_sales'],
                    black_status,
                    data['sample_order'],
                    data['sku_name'],
                    video_status,
                    data['remark']
                ))
        else:
            # 不合并，直接显示所有记录
            for row in self.c.fetchall():
                # 确保有9个值
                row = list(row)
                while len(row) < 9:
                    row.append('' if len(row) < 6 else 0)
                
                daren_id, name, shop, sales, is_black, sample_order, sku_name, has_video, remark = row
                black_status = "是" if is_black else "否"
                video_status = "是" if has_video else "否"
                self.tree.insert("", tk.END, values=(
                    daren_id,
                    name,
                    shop,
                    sales,
                    black_status,
                    sample_order,
                    sku_name,
                    video_status,
                    remark
                ))                                                                  
  
    def delete_selected(self):
        """删除选中的记录"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("警告", "请先选择要删除的记录")
            return
        
        if not messagebox.askyesno("确认删除", f"确定要删除选中的 {len(selected_items)} 条记录吗%s,"):
            return
        
        try:
            deleted_count = 0
            for item in selected_items:
                item_id = self.tree.item(item)['values'][0]
                self.c.execute("DELETE FROM daren WHERE id=%s,", (item_id,))
                deleted_count += 1
            
            self.conn.commit()
            messagebox.showinfo("删除成功", f"已成功删除 {deleted_count} 条记录")
            self.search_daren()  # 刷新显示
        
        except Exception as e:
            messagebox.showerror("删除失败", f"删除过程中出错:\n{str(e)}")
    
    def export_to_excel(self):
        """导出数据到Excel"""
        export_dialog = tk.Toplevel(self.root)
        export_dialog.title("导出选项")
        export_dialog.geometry("300x300")
        export_dialog.grab_set()
        
        ttk.Label(export_dialog, text="导出选项", font=('Arial', 12)).pack(pady=10)
        
        # 导出范围选择
        ttk.Label(export_dialog, text="导出范围:").pack(anchor="w", padx=20)
        range_frame = ttk.Frame(export_dialog)
        range_frame.pack(fill="x", padx=20)
        
        self.export_all = tk.BooleanVar(value=True)
        ttk.Radiobutton(range_frame, text="全部数据", variable=self.export_all, value=True).pack(side="left")
        ttk.Radiobutton(range_frame, text="当前查询结果", variable=self.export_all, value=False).pack(side="left", padx=10)
        
        # 列选择
        ttk.Label(export_dialog, text="选择要导出的列:").pack(anchor="w", padx=20, pady=(10,0))
        
        columns_frame = ttk.Frame(export_dialog)
        columns_frame.pack(fill="x", padx=20)
        
        self.export_columns = {
            "id": tk.BooleanVar(value=True),
            "name": tk.BooleanVar(value=True),
            "shops": tk.BooleanVar(value=True),
            "sales": tk.BooleanVar(value=True),
            "black": tk.BooleanVar(value=True),
            "sample_order": tk.BooleanVar(value=True),
            "sku_name": tk.BooleanVar(value=True),
            "has_video": tk.BooleanVar(value=True),
            "remark": tk.BooleanVar(value=True)
        }
        
        ttk.Checkbutton(columns_frame, text="ID", variable=self.export_columns["id"]).grid(row=0, column=0, sticky="w")
        ttk.Checkbutton(columns_frame, text="达人名称", variable=self.export_columns["name"]).grid(row=0, column=1, sticky="w")
        ttk.Checkbutton(columns_frame, text="所属店铺", variable=self.export_columns["shops"]).grid(row=1, column=0, sticky="w")
        ttk.Checkbutton(columns_frame, text="销售额", variable=self.export_columns["sales"]).grid(row=1, column=1, sticky="w")
        ttk.Checkbutton(columns_frame, text="黑名单", variable=self.export_columns["black"]).grid(row=2, column=0, sticky="w")
        ttk.Checkbutton(columns_frame, text="样品订单号", variable=self.export_columns["sample_order"]).grid(row=2, column=1, sticky="w")
        ttk.Checkbutton(columns_frame, text="产品 SKU 名称", variable=self.export_columns["sku_name"]).grid(row=3, column=0, sticky="w")
        ttk.Checkbutton(columns_frame, text="是否创作视频", variable=self.export_columns["has_video"]).grid(row=3, column=1, sticky="w")
        ttk.Checkbutton(columns_frame, text="备注", variable=self.export_columns["remark"]).grid(row=4, column=0, sticky="w")
        
        # 操作按钮
        button_frame = ttk.Frame(export_dialog)
        button_frame.pack(pady=20)
        
        ttk.Button(button_frame, text="导出", command=lambda: self._perform_export(export_dialog)).pack(side="left", padx=10)
        ttk.Button(button_frame, text="取消", command=export_dialog.destroy).pack(side="left")
    
    def _perform_export(self, dialog):
        """执行实际的导出操作"""
        try:
            default_name = f"达人数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx")],
                title="保存为Excel文件",
                initialfile=default_name
            )
            
            if not filepath:
                return
            
            wb = Workbook()
            ws_main = wb.active
            ws_main.title = "达人数据"
            
            # 准备列头
            columns = []
            if self.export_columns["id"].get():
                columns.append(("ID", "id"))
            if self.export_columns["name"].get():
                columns.append(("达人名称", "name"))
            if self.export_columns["shops"].get():
                columns.append(("所属店铺", "shops"))
            if self.export_columns["sales"].get():
                columns.append(("销售额", "sales"))
            if self.export_columns["black"].get():
                columns.append(("是否黑名单", "black"))
            if self.export_columns["sample_order"].get():
                columns.append(("样品订单号", "sample_order"))
            if self.export_columns["sku_name"].get():
                columns.append(("产品 SKU 名称", "sku_name"))
            if self.export_columns["has_video"].get():
                columns.append(("是否创作视频", "has_video"))
            if self.export_columns["remark"].get():
                columns.append(("备注", "remark"))

            ws_main.append([col[0] for col in columns])
            
            # 获取数据
            if self.export_all.get():
                self.c.execute("SELECT * FROM daren")
            else:
                name = self.search_entry.get()
                if name:
                    self.c.execute("SELECT * FROM daren WHERE name LIKE %s,", ('%'+name+'%',))
                else:
                    self.c.execute("SELECT * FROM daren")
            
            # 添加数据
            for row in self.c.fetchall():
                data = []
                for col in columns:
                    col_name = col[1]
                    if col_name == "has_video":
                        data.append("是" if row[7] else "否")
                    elif col_name == "remark":
                        data.append(row[8])  # 确保索引正确
                    else:
                        data.append(row[columns.index(col)])
                ws_main.append(data)
            
            # 调整列宽...
            wb.save(filepath)
            messagebox.showinfo("导出成功", f"数据已成功导出到:\n{filepath}")
            dialog.destroy()
        
        except Exception as e:
            messagebox.showerror("导出失败", f"导出过程中出错:\n{str(e)}")
    
    def import_data(self):
        """导入Excel或CSV文件数据"""
        filepath = filedialog.askopenfilename(
            filetypes=[
                ("Excel文件", "*.xlsx"),
                ("CSV文件", "*.csv"),
                ("所有文件", "*.*")
            ],
            title="选择要导入的文件"
        )
        
        if not filepath:
            return
        
        try:
            if filepath.endswith('.csv'):
                df = pd.read_csv(filepath)
            else:
                df = pd.read_excel(filepath, engine='openpyxl')
            
            # 检查必要列
            if '达人名称' not in df.columns:
                raise ValueError("文件必须包含'达人名称'列")
            
            # 显示预览
            self.show_import_preview(df)
        
        except Exception as e:
            messagebox.showerror("导入失败", f"读取文件出错:\n{str(e)}")
    
    def show_import_preview(self, df):
        """显示导入数据预览"""
        preview_dialog = tk.Toplevel(self.root)
        preview_dialog.title("导入预览")
        preview_dialog.geometry("800x600")
        preview_dialog.grab_set()
        
        # 创建Treeview显示预览
        tree = ttk.Treeview(preview_dialog)
        tree["columns"] = list(df.columns)
        
        for col in df.columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        
        # 添加数据
        for _, row in df.iterrows():
            tree.insert("", tk.END, values=list(row))
        
        tree.pack(fill="both", expand=True)
        
        # 添加操作按钮
        btn_frame = ttk.Frame(preview_dialog)
        btn_frame.pack(pady=10)
        
        ttk.Button(btn_frame, text="确认导入", 
                 command=lambda: self._confirm_import(df, preview_dialog)).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="取消", 
                 command=preview_dialog.destroy).pack(side="left")
    
    def _confirm_import(self, df, dialog):
        """确认导入预览的数据"""
        try:
            total = 0
            success = 0
            
            for _, row in df.iterrows():
                try:
                    # 处理黑名单字段
                    black_status = str(row.get('黑名单', '')).lower()
                    is_black = 1 if black_status in ('是', 'yes', 'true', '1') else 0
                    
                    # 处理视频字段
                    video_status = str(row.get('是否创作视频', '')).lower()
                    has_video = 1 if video_status in ('是', 'yes', 'true', '1') else 0
                    
                    data = {
                        'name': str(row.get('达人名称', '')),
                        'shop': str(row.get('店铺名称', '')),
                        'sales': float(row.get('销售额', 0)),
                        'is_black': is_black,
                        'sample_order': str(row.get('样品订单号', '')),
                        'sku_name': str(row.get('产品 SKU 名称', '')),
                        'has_video': has_video,
                        'remark': str(row.get('备注', ''))
                    }
                    
                    self.c.execute(
                        """INSERT INTO daren 
                           (name, shop, sales, is_black, sample_order, sku_name, has_video, remark)
                           VALUES (:name, :shop, :sales, :is_black, :sample_order, :sku_name, :has_video, :remark)""",
                        data
                    )
                    success += 1
                except Exception as e:
                    print(f"跳过一行数据，错误: {str(e)}")
                total += 1
            
            self.conn.commit()
            dialog.destroy()
            
            messagebox.showinfo(
                "导入完成",
                f"成功导入 {success}/{total} 条记录\n"
                f"失败 {total-success} 条\n"
                "失败记录请检查数据格式"
            )
            
            self.search_daren()
        
        except Exception as e:
            messagebox.showerror("导入失败", f"导入过程中出错:\n{str(e)}")

    def export_template(self):
        """导出导入模板文件"""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel模板", "*.xlsx")],
            title="保存导入模板",
            initialfile="达人数据导入模板.xlsx"
        )
        
        if not filepath:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "达人数据"
            
            headers = ["达人名称", "店铺名称", "销售额", "黑名单", "样品订单号", "产品 SKU 名称", "是否创作视频", "备注"]
            examples = ["张三", "张三的店铺", "5000.00", "是", "123456", "SKU123", "是", "这是备注"]
            
            ws.append(headers)
            ws.append(examples)
            
            for cell in ws[1]:
                cell.style = 'Headline 2'
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            
            wb.save(filepath)
            messagebox.showinfo("模板导出", f"模板已保存到:\n{filepath}")
        
        except Exception as e:
            messagebox.showerror("导出失败", f"模板导出失败:\n{str(e)}")
    
    def __del__(self):
        """析构函数，关闭数据库连接"""
        self.conn.close()

if __name__ == "__main__":
    # 初始化数据库
    conn = sqlite3.connect('daren.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS talent (
                 id INT AUTO_INCREMENT PRIMARY KEY,
                 name VARCHAR(255) NOT NULL,
                 shop VARCHAR(255),
                 sales FLOAT,
                 is_black BOOLEAN DEFAULT 0,
                 sample_order VARCHAR(255),
                 sku_name VARCHAR(255),
                 has_video BOOLEAN DEFAULT 0,
                 remark TEXT
    )''')  # 确保有9个字段
    conn.commit()
    conn.close()
    

    
    # 启动应用
    root = tk.Tk()
    app = DarenApp(root)
    root.mainloop()
